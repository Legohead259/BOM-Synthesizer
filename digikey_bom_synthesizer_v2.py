import csv
from openpyxl import load_workbook

# =====CONSTANTS=====

SUPPLIER_COLUMN = 'C'
SUPPLIER_PART_NUM_COLUMN = 'D'
QUANTITY_COLUMN = 'I'
INV_QUANTITY_COLUMN = "H"

bom_dict = {}  # Key: BOM Worksheet, Value: Quantity
component_dict = {}  # Key: Digikey part number, Value: Quantity 

"""
Loads the worksheets specified in inputs.csv into the BOM list for processing
"""
def load_worksheets():
    with open("inputs.csv") as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        line_count = 0
        for row in csv_reader:
            # print(row)  # Debug
            if line_count != 0:
                try:
                    _ws_temp = load_workbook("BOMs/"+row[0])[row[1]]  # Load worksheet into memory
                    if _ws_temp not in bom_dict:  # Check if worksheet has been loaded into the BOM dictionary
                        bom_dict[_ws_temp] = row[2]  # Adds BOM quantity to dictionary value
                except (FileNotFoundError):  # BOM not found in the BOMs folder
                    print(f'"{row[0]}" not found. Try placing the workbook in "BOMs" folder or check file name!')
                except (KeyError):  # Specified worksheet name not found in specified workbook
                    print(f'Sheet "{row[1]}" not found in workbook "{row[0]}". Please check the sheet name in the workbook!')
            line_count += 1
    # print(bom_list)  # Debug


"""
Parses the DigiKey components found in the passed sheet into the component dictionary
"""
def parse_to_component_dictionary(sheet):
    _row_number = 0
    for supplier in sheet[SUPPLIER_COLUMN]:
        if supplier.value == "Digikey":  # Check if component in row is supplied by Digikey
            _supplier_pn_temp = sheet[SUPPLIER_PART_NUM_COLUMN][_row_number].value

            if _supplier_pn_temp not in component_dict:  # Check if component is not already in the component dictionary
                # print(f'{_supplier_pn_temp} found in dictionary!')  # Debug
                component_dict[_supplier_pn_temp] = sheet[QUANTITY_COLUMN][_row_number].value * int(bom_dict.get(sheet))  # Add component and quantity (times board quantity) to dictionary
            else:  # Component found in dictionary already
                _prev_value = component_dict.get(_supplier_pn_temp)
                component_dict[_supplier_pn_temp] = sheet[QUANTITY_COLUMN][_row_number].value * int(bom_dict.get(sheet)) + _prev_value  # Update component quantity
        _row_number += 1
    # print(component_dict)  # Debug


"""
Reads an inventory work book and subtracts stocked components from the component dictionary
"""
def remove_stocked_components(inv_fp):
    _MINIMUM_STOCK_QTY = 5
    _inventory_ws = load_workbook(inv_fp).active

    _row_number = 0
    for part in _inventory_ws[SUPPLIER_PART_NUM_COLUMN]:
        if part.value in component_dict:
            print(f'Part "{part.value}" found in component dictionary!')  # Debug
            _adj_value = _inventory_ws[INV_QUANTITY_COLUMN][_row_number].value - component_dict.get(part.value)

            if _adj_value < 0:
                _adj_value += (0-_adj_value)  # if the adjusted order value is less than 0, add enough components to zero it out

            if (_adj_value) < _MINIMUM_STOCK_QTY:
                _adj_value += _MINIMUM_STOCK_QTY - _adj_value

            component_dict[part.value] = _adj_value
        _row_number += 1
    # print(component_dict)  # Debug

def generate_order_list():
    try:
        with open("digikey-order.csv", 'w') as csvfile:
            for component in component_dict.keys():
                csvfile.write("%s,%s\n" % (component, component_dict[component]))
    except (IOError):
        print("I/O Error!")

if __name__ == "__main__":
    load_worksheets()
    for ws in bom_dict:
        parse_to_component_dictionary(ws)
    remove_stocked_components("inventory.xlsx")
    generate_order_list()